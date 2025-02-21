import csv
import json
from datetime import datetime
from typing import List, Dict, Type

import openpyxl
from fastapi import HTTPException
from pydantic import BaseModel
from sqlalchemy import DateTime
from sqlalchemy.orm import Session


class DataExporter:
    """数据导出工具类"""

    @staticmethod
    def export_to_csv(data: List[Dict], fields: List[str], filename: str) -> str:
        """导出数据到CSV文件"""
        try:
            with open(filename, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.DictWriter(f, fieldnames=fields)
                writer.writeheader()
                writer.writerows(data)
            return filename
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Failed to export CSV: {str(e)}")

    @staticmethod
    def export_to_excel(data: List[Dict], fields: List[str], filename: str) -> str:
        """导出数据到Excel文件"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active

            # 写入表头
            for col, field in enumerate(fields, 1):
                ws.cell(row=1, column=col, value=field)

            # 写入数据
            for row, item in enumerate(data, 2):
                for col, field in enumerate(fields, 1):
                    ws.cell(row=row, column=col, value=item.get(field))

            wb.save(filename)
            return filename
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Failed to export Excel: {str(e)}")

    @staticmethod
    def export_to_json(data: List[Dict], filename: str) -> str:
        """导出数据到JSON文件"""
        try:
            with open(filename, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            return filename
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Failed to export JSON: {str(e)}")


class DataImporter:
    """数据导入工具类"""

    @staticmethod
    def validate_data(data: Dict, model_class: Type[BaseModel]) -> bool:
        """验证数据格式"""
        required_fields = [
            column.name
            for column in model_class.__table__.columns
            if not column.nullable and column.name not in ["id", "create_time", "update_time", "delete_time"]
        ]
        return all(field in data for field in required_fields)

    @staticmethod
    def import_from_csv(filename: str, model_class: Type[BaseModel], db: Session) -> int:
        """从CSV文件导入数据"""
        try:
            imported_count = 0
            with open(filename, "r", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    if DataImporter.validate_data(row, model_class):
                        # 处理日期时间字段
                        for key, value in row.items():
                            if isinstance(getattr(model_class, key).type, DateTime):
                                row[key] = datetime.fromisoformat(value) if value else None

                        db_obj = model_class(**row)
                        db.add(db_obj)
                        imported_count += 1

                db.commit()
            return imported_count
        except Exception as e:
            db.rollback()
            raise HTTPException(status_code=500, detail=f"Failed to import CSV: {str(e)}")

    @staticmethod
    def import_from_excel(filename: str, model_class: Type[BaseModel], db: Session) -> int:
        """从Excel文件导入数据"""
        try:
            imported_count = 0
            wb = openpyxl.load_workbook(filename)
            ws = wb.active

            # 获取表头
            headers = [cell.value for cell in ws[1]]

            # 读取数据
            for row in ws.iter_rows(min_row=2):
                data = {headers[i]: cell.value for i, cell in enumerate(row)}
                if DataImporter.validate_data(data, model_class):
                    # 处理日期时间字段
                    for key, value in data.items():
                        if isinstance(getattr(model_class, key).type, DateTime):
                            data[key] = datetime.fromisoformat(str(value)) if value else None

                    db_obj = model_class(**data)
                    db.add(db_obj)
                    imported_count += 1

            db.commit()
            return imported_count
        except Exception as e:
            db.rollback()
            raise HTTPException(status_code=500, detail=f"Failed to import Excel: {str(e)}")

    @staticmethod
    def import_from_json(filename: str, model_class: Type[BaseModel], db: Session) -> int:
        """从JSON文件导入数据"""
        try:
            imported_count = 0
            with open(filename, "r", encoding="utf-8") as f:
                data_list = json.load(f)

            for data in data_list:
                if DataImporter.validate_data(data, model_class):
                    # 处理日期时间字段
                    for key, value in data.items():
                        if isinstance(getattr(model_class, key).type, DateTime):
                            data[key] = datetime.fromisoformat(value) if value else None

                    db_obj = model_class(**data)
                    db.add(db_obj)
                    imported_count += 1

            db.commit()
            return imported_count
        except Exception as e:
            db.rollback()
            raise HTTPException(status_code=500, detail=f"Failed to import JSON: {str(e)}")
